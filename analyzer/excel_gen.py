from openpyxl import Workbook
wb1 = Workbook()
n=16
names=["bi_gram_dist","bi_gram_cluster","bisorgo","chandrabindu","char_tri_gram","char_tri_gram_with_example","cluster_dist","coda_dist","cv_dist","cv_file","dipthong_dist","mono_graph","onset_dist","phone_dist","word_freq","word_ipa_freq","word_ipa_freq_v1","prefix_dist"]
for i in range(n):
    ws = wb1.create_sheet(names[i])

    #code on formatting sheet, optimization problem

wb1.save('output/output.xlsx')

import pandas as pd

#read_file = pd.read_csv (r'../a.txt')
#read_file.to_csv (r'a.csv', index=None)


from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
#wb = load_workbook('outfile.xlsx')  # load as openpyxl workbook; useful to keep the original layout
                                # which is discarded in the following dataframe
#for i in range(n):		                         
#   print(names[i])
#   writer = pd.ExcelWriter('output.xlsx')
#   df = pd.read_csv(names[i]+'.csv')
#   #df.to_excel(writer,sheet_name=names[i],index=False)
#   with pd.ExcelWriter('output.xlsx') as writer:
#    df.to_excel(writer, sheet_name=names[i])
excelBook = load_workbook('output/output.xlsx')

with pd.ExcelWriter('output/output.xlsx', engine='openpyxl') as writer:
    # Save your file workbook as base
    writer.book = excelBook
    writer.sheets = dict((ws.title, ws) for ws in excelBook.worksheets)

    # Now here add your new sheets
    for i in range(n):
        df = pd.read_csv("output/"+names[i]+'.csv')
        df.to_excel(writer,names[i], index = False)

    # Save the file
    writer.save()

#writer.save()
#for i in range(n):		                         
#    df = pd.read_csv(names[i]+'.csv')  # load as dataframe (modifications will be easier with pandas API!)
#    ws = wb[names[i]]
#    print(names[i])
#	#df.iloc[1, 1] = 'hello world'    # modify a few things
#    rows = dataframe_to_rows(df, index=False)
#    count=0
#    for r_idx, row in enumerate(rows, 1):
#        for c_idx, value in enumerate(row, 1):
#            count=count+1
#            if count%200==0:
#               print(count)
#            ws.cell(row=r_idx, column=c_idx, value=value)
#        wb.save('outfile.xlsx')


